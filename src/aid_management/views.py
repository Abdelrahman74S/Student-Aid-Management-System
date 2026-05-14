from django.core.cache import cache
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from core.mixins import EnhancedListMixin
from .filters import AidApplicationFilter
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.generic import (
    FormView, ListView, CreateView, UpdateView, 
    DetailView, TemplateView
)
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Avg, Count, Q
from django.core.exceptions import ValidationError
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from django.conf import settings

from .models import (
    SupportCycle, AidApplication, ScoringRule, 
    CommitteeReview, BudgetAllocation
)
from .forms import StudentApplicationForm, CommitteeReviewForm, ApplicationDocumentFormSet
from accounts.models import User, StudentProfile, ReviewerProfile
from accounts.mixins import (
    StudentRequiredMixin, ReviewerRequiredMixin, 
    CommitteeHeadRequiredMixin
)

# =================================================================
# 1. Student Portal – محمود
# =================================================================

class StudentApplicationListView(LoginRequiredMixin, StudentRequiredMixin, EnhancedListMixin, ListView):
    model = AidApplication
    template_name = "aid_management/student/application_list.html"
    context_object_name = "applications"
    filterset_class = AidApplicationFilter
    search_fields = ['serial_number', 'cycle__name']
    ordering_fields = ['created_at', 'status']
    default_ordering = '-created_at'

    def get_queryset(self):
        queryset = super().get_queryset().filter(
            student=self.request.user.student_profile,
            deleted_at__isnull=True
        ).select_related('cycle')
        return queryset


class StudentApplicationCreateView(LoginRequiredMixin, StudentRequiredMixin, CreateView):
    model = AidApplication
    form_class = StudentApplicationForm
    template_name = "aid_management/student/application_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.active_cycle = SupportCycle.objects.filter(status='OPEN').first()
        if not self.active_cycle or not self.active_cycle.is_open_for_application:
            messages.error(request, "لا توجد دورة دعم متاحة للتقديم حالياً.")
            return redirect('aid_management:application_list')

        if AidApplication.objects.filter(
            student=request.user.student_profile,
            cycle=self.active_cycle
        ).exists():
            messages.warning(request, "لقد قمت بالتقديم بالفعل في هذه الدورة.")
            return redirect('aid_management:application_list')

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = ApplicationDocumentFormSet(self.request.POST, self.request.FILES, instance=self.object)
        else:
            context['formset'] = ApplicationDocumentFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        form.instance.student = self.request.user.student_profile
        form.instance.cycle = self.active_cycle
        
        if form.is_valid() and formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            messages.success(self.request, "تم حفظ المسودة بنجاح.")
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('aid_management:application_list')


class StudentApplicationUpdateView(LoginRequiredMixin, StudentRequiredMixin, UpdateView):
    model = AidApplication
    form_class = StudentApplicationForm
    template_name = "aid_management/student/application_form.html"

    def get_queryset(self):
        return AidApplication.objects.filter(
            student=self.request.user.student_profile,
            status='DRAFT'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = ApplicationDocumentFormSet(self.request.POST, self.request.FILES, instance=self.object)
        else:
            context['formset'] = ApplicationDocumentFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if form.is_valid() and formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            messages.success(self.request, "تم تحديث المسودة بنجاح.")
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('aid_management:application_list')

class StudentApplicationSubmitView(LoginRequiredMixin, StudentRequiredMixin, View):
    def post(self, request, pk):
        application = get_object_or_404(
            AidApplication,
            id=pk,
            student=request.user.student_profile
        )

        try:
            with transaction.atomic():
                application.freeze_student_data()
                application.submit(request=request)

            messages.success(request, "تم إرسال الطلب بنجاح.")
        except ValidationError as e:
            messages.error(request, str(e))

        return redirect('aid_management:application_detail', pk=pk)


class StudentApplicationWithdrawView(LoginRequiredMixin, StudentRequiredMixin, View):
    def post(self, request, pk):
        application = get_object_or_404(
            AidApplication,
            id=pk,
            student=request.user.student_profile
        )

        try:
            application.soft_delete()
            messages.info(request, "تم سحب الطلب بنجاح.")
        except ValidationError as e:
            messages.error(request, str(e))

        return redirect('aid_management:application_list')


class StudentApplicationDetailView(LoginRequiredMixin, StudentRequiredMixin, DetailView):
    model = AidApplication
    template_name = "aid_management/student/application_detail.html"
    context_object_name = "application"

    def get_queryset(self):
        return AidApplication.objects.filter(
            student=self.request.user.student_profile
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        application = self.get_object()

        if application.status in ['SCORED', 'APPROVED', 'REJECTED', 'DISBURSED']:
            context['reviews_summary'] = application.reviews.filter(
                is_submitted=True
            ).aggregate(
                avg_total=Avg('total_score'),
                count=Count('id')
            )

        if application.status in ['APPROVED', 'DISBURSED']:
            context['allocation'] = application.allocations.filter(
                status__in=['PENDING', 'DISBURSED']
            ).first()

        context['timeline'] = self.get_application_timeline(application)
        return context

    def get_application_timeline(self, app):
        events = []
        events.append({'label': 'إنشاء المسودة', 'date': app.created_at, 'status': 'completed'})

        if app.submission_date:
            events.append({'label': 'تم تقديم الطلب', 'date': app.submission_date, 'status': 'completed'})

        if app.status == 'UNDER_REVIEW':
            events.append({'label': 'قيد المراجعة', 'date': None, 'status': 'active'})

        if app.decision_date:
            label = 'تم القبول' if app.status in ['APPROVED', 'DISBURSED'] else 'تم الرفض'
            events.append({'label': label, 'date': app.decision_date, 'status': 'completed'})

        return events


# =================================================================
# 2. Reviewer System – ياسمين
# =================================================================

class ReviewerTaskListView(LoginRequiredMixin, ReviewerRequiredMixin, ListView):
    template_name = "aid_management/reviewer/task_list.html"
    context_object_name = "reviews"

    def get_queryset(self):
        return CommitteeReview.objects.filter(
            reviewer=self.request.user,
            is_submitted=False
        ).select_related('application__student__user', 'application__cycle')


class ApplicationScoringView(LoginRequiredMixin, ReviewerRequiredMixin, UpdateView):
    model = CommitteeReview
    form_class = CommitteeReviewForm
    template_name = "aid_management/reviewer/scoring.html"

    def get_queryset(self):
        return CommitteeReview.objects.filter(reviewer=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['application'] = self.get_object().application
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        application = self.get_object().application

        # بيانات التقييم التلقائي
        context['auto_breakdown'] = application.auto_score_breakdown
        context['auto_score'] = application.auto_score

        # حساب الاقتراحات الحية (في حال عدم وجود breakdown مخزّن)
        if not application.auto_score_breakdown:
            from .services import get_application_scoring_details
            details = get_application_scoring_details(application)
            context['auto_breakdown'] = details
            context['auto_score'] = details['total_auto_score']

        context['application'] = application

        # بيانات البحث الاجتماعي إن وُجدت
        try:
            context['social_research'] = application.social_research
        except Exception:
            context['social_research'] = None

        return context

    def form_valid(self, form):
        review = form.save(commit=False)

        if 'submit_final' in self.request.POST:
            try:
                review.submit()
                messages.success(self.request, "تم اعتماد التقييم نهائياً.")
            except ValidationError as e:
                messages.error(self.request, str(e))
                return self.form_invalid(form)
        else:
            review.save()
            messages.info(self.request, "تم حفظ كمسودة.")

        return redirect('aid_management:reviewer_task_list')


class ReviewConflictFlagView(LoginRequiredMixin, ReviewerRequiredMixin, View):
    def post(self, request, pk):
        review = get_object_or_404(CommitteeReview, id=pk, reviewer=request.user)
        review.conflict_of_interest = True
        review.qualitative_notes = request.POST.get('reason', 'تعارض مصالح')
        review.save()
        messages.warning(request, "تم الإبلاغ عن تعارض مصالح.")
        return redirect('aid_management:reviewer_task_list')


# =================================================================
# 3. Committee – كريم
# =================================================================

class CommitteeHeadDashboardView(LoginRequiredMixin, CommitteeHeadRequiredMixin, TemplateView):
    template_name = "aid_management/committee/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_id = self.request.user.id
        cache_key = f"committee_head_dashboard_{user_id}"
        
        cached_data = cache.get(cache_key)
        if cached_data:
            context.update(cached_data)
            return context

        cycle = SupportCycle.objects.filter(status__in=['OPEN', 'UNDER_REVIEW']).first()

        if cycle:
            # Expensive aggregations
            total_apps = cycle.applications.count()
            submitted_apps = cycle.applications.exclude(status='DRAFT').count()
            reviews_submitted = CommitteeReview.objects.filter(
                application__cycle=cycle, is_submitted=True
            ).count()
            
            # Need to know total reviews required (e.g. 2 per app)
            total_needed = total_apps * 2 
            
            # أعلى الطلبات من حيث التقييم اليدوي (SCORED)
            top_apps = cycle.applications.filter(status='SCORED').annotate(
                avg_score=Avg('reviews__total_score')
            ).order_by('-avg_score')[:5]

            # الترتيب الأولي التلقائي — للطلبات المقدمة قبل المراجعة اليدوية
            priority_ranking = cycle.applications.filter(
                status__in=['SUBMITTED', 'UNDER_REVIEW'],
                auto_score__gt=0
            ).select_related(
                'student__user', 'student__program'
            ).order_by('-auto_score')[:10]

            dashboard_data = {
                'active_cycle': cycle,
                'total_apps_count': total_apps,
                'submitted_apps_count': submitted_apps,
                'reviews_completed': reviews_submitted,
                'total_reviews_needed': total_needed,
                'budget_reserved_percent': (cycle.reserved_budget / cycle.total_budget * 100 
                                            if cycle.total_budget else 0),
                'budget_disbursed_percent': (cycle.disbursed_budget / cycle.total_budget * 100 
                                            if cycle.total_budget else 0),
                'top_applications': list(top_apps),
                'priority_ranking': list(priority_ranking),
            }
            cache.set(cache_key, dashboard_data, timeout=settings.CACHE_TTL)
            context.update(dashboard_data)
            
        return context


@method_decorator(cache_page(60 * 15, key_prefix='application_ranking_list'), name='dispatch')
class ApplicationRankListView(LoginRequiredMixin, CommitteeHeadRequiredMixin, EnhancedListMixin, ListView):
    template_name = "aid_management/committee/ranking_list.html"
    context_object_name = "ranked_applications"
    filterset_class = AidApplicationFilter
    search_fields = ['serial_number', 'student__user__first_name', 'student__user__last_name']
    ordering_fields = ['avg_score', 'created_at']
    default_ordering = '-avg_score'

    def get_queryset(self):
        cycle_id = self.request.GET.get('cycle')
        managed_programs = self.request.user.committee_head_profile.managed_programs.all()
        qs = super().get_queryset().filter(status__in=['SCORED', 'APPROVED', 'DISBURSED'], student__program__in=managed_programs)
        if cycle_id:
            qs = qs.filter(cycle_id=cycle_id)
        return qs.annotate(avg_score=Avg('reviews__total_score')).order_by('-avg_score')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cycle_id = self.request.GET.get('cycle')
        if cycle_id:
            context['active_cycle'] = get_object_or_404(SupportCycle, id=cycle_id)
        else:
            context['active_cycle'] = SupportCycle.objects.filter(status__in=['OPEN', 'UNDER_REVIEW']).first()
        return context


class FinalDecisionUpdateView(LoginRequiredMixin, CommitteeHeadRequiredMixin, View):
    def post(self, request, pk):
        managed_programs = request.user.committee_head_profile.managed_programs.all()
        application = get_object_or_404(AidApplication, id=pk, student__program__in=managed_programs)
        decision = request.POST.get('decision')
        amount = request.POST.get('amount', 0)

        try:
            with transaction.atomic():
                application.status = decision
                application.committee_decision = request.POST.get('notes')
                application.decision_date = timezone.now()
                application.save()

                if decision == 'APPROVED':
                    BudgetAllocation.objects.create(
                        cycle=application.cycle,
                        application=application,
                        amount_allocated=amount
                    )
            messages.success(request, f"تم تحديث القرار للطلب {application.serial_number}")
        except Exception as e:
            messages.error(request, f"خطأ: {str(e)}")
        return redirect('aid_management:application_rank_list')


class AutomatedDistributionView(LoginRequiredMixin, CommitteeHeadRequiredMixin, View):
    def post(self, request):
        cycle = SupportCycle.objects.filter(status='UNDER_REVIEW').first()
        reviewers = User.objects.filter(groups__name='Reviewers')
        applications = AidApplication.objects.filter(cycle=cycle, status='SUBMITTED')

        created_count = 0
        with transaction.atomic():
            for app in applications:
                for rev in reviewers[:2]:
                    CommitteeReview.objects.get_or_create(application=app, reviewer=rev)
                app.status = 'UNDER_REVIEW'
                app.save()
                created_count += 1
        messages.success(request, f"تم توزيع {created_count} طلب على المراجعين.")
        return redirect('aid_management:committee_dashboard')


class CycleStatusTransitionView(LoginRequiredMixin, CommitteeHeadRequiredMixin, View):
    def post(self, request, pk):
        cycle = get_object_or_404(SupportCycle, id=pk)
        new_status = request.POST.get('status')
        valid_statuses = [c[0] for c in SupportCycle.STATUS_CHOICES]

        if new_status not in valid_statuses:
            messages.error(request, "حالة غير صالحة.")
            return redirect('aid_management:committee_dashboard')

        cycle_flow = {
            'DRAFT': ['OPEN'],
            'OPEN': ['UNDER_REVIEW', 'CLOSED'],
            'UNDER_REVIEW': ['CLOSED'],
            'CLOSED': ['ARCHIVED']
        }

        if new_status not in cycle_flow.get(cycle.status, []):
            messages.error(request, "انتقال غير منطقي.")
            return redirect('aid_management:committee_dashboard')

        try:
            with transaction.atomic():
                if new_status == 'OPEN':
                    if not cycle.rules.exists():
                        raise ValidationError("يجب إضافة قواعد التقييم أولاً.")
                    cycle.scoring_rules_snapshot = list(
                        cycle.rules.values('criteria_type', 'points', 'weight')
                    )
                elif new_status == 'CLOSED':
                    cycle.is_locked = True

                cycle.status = new_status
                cycle.save()
            messages.success(request, f"تم تغيير الحالة إلى {cycle.get_status_display()}")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, str(e))
        return redirect('aid_management:committee_dashboard')


class DisbursementExportView(LoginRequiredMixin, CommitteeHeadRequiredMixin, View):
    def get(self, request, cycle_id):
        cycle = get_object_or_404(SupportCycle, id=cycle_id)
        export_type = request.GET.get('format', 'excel')
        
        # Get all approved or disbursed applications for this cycle
        applications = AidApplication.objects.filter(
            cycle=cycle, 
            status__in=['APPROVED', 'DISBURSED']
        ).select_related('student__user', 'budget_allocation')

        if export_type == 'excel':
            return self.export_excel(cycle, applications)
        elif export_type == 'pdf':
            return self.export_pdf(cycle, applications)
        else:
            return HttpResponse("Unsupported format", status=400)

    def export_excel(self, cycle, applications):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "كشف الصرف المالي"
        ws.sheet_view.rightToLeft = True

        # Headers
        headers = [
            'الرقم المتسلسل', 'اسم الطالب', 'الرقم القومي', 'الرقم الجامعي', 
            'المبلغ المخصص', 'اسم البنك', 'رقم الحساب (IBAN)', 
            'مزود المحفظة', 'رقم المحفظة'
        ]
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row_num, app in enumerate(applications, 2):
            # Check for budget allocation
            allocation = BudgetAllocation.objects.filter(application=app).first()
            amount = allocation.amount_allocated if allocation else 0
            
            ws.cell(row=row_num, column=1, value=app.serial_number)
            ws.cell(row=row_num, column=2, value=app.student.user.get_full_name())
            ws.cell(row=row_num, column=3, value=app.student.user.national_id)
            ws.cell(row=row_num, column=4, value=app.student.student_id)
            ws.cell(row=row_num, column=5, value=amount)
            ws.cell(row=row_num, column=6, value=app.student.bank_name or "N/A")
            ws.cell(row=row_num, column=7, value=app.student.bank_account_number or "N/A")
            ws.cell(row=row_num, column=8, value=app.student.wallet_provider or "N/A")
            ws.cell(row=row_num, column=9, value=app.student.wallet_number or "N/A")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="disbursement_{cycle.serial_number}.xlsx"'
        wb.save(response)
        return response

    def export_pdf(self, cycle, applications):
        from assets_reporting.utils import render_to_pdf
        context = {
            'cycle': cycle,
            'applications': applications,
            'today': timezone.now(),
        }
        pdf = render_to_pdf('aid_management/exports/disbursement_list.html', context)
        if pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="disbursement_{cycle.serial_number}.pdf"'
            return response
        return HttpResponse("Error generating PDF", status=500)